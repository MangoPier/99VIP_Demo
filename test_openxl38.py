# 99 VIP 38
import openpyxl
import unittest
from ddt import data,ddt,unpack

def read_excel():
    # 加载工作簿
    excel = openpyxl.load_workbook("test_data.xlsx")
    # 指定表单
    sheet1 = excel["Sheet1"]
    max_row1 = sheet1.max_row
    print(max_row1)
    max_column1 = sheet1.max_column
    print(max_column1)

    # 获取表格里面的数据存储在列表
    # 用到嵌套循环遍历表格的数据
    list_datas = []
    for row in range(2,max_row1+1):
        # 定义一个列表来存储每一行的数据
        rowlist = []
        for col in range(1,max_column1+1):
            # 列表添加数据
            rowlist.append(sheet1.cell(row,col).value)
        list_datas.append(rowlist)
    # print(list_datas)
    return list_datas


# read_excel()

@ddt
class TestLogin(unittest.TestCase):
    @data(*read_excel()) # * 代表不定个数的形参
    @unpack
    def test_case(self,id,username,pwd):
        print(f"这是第{id}条测试用例")
        self.assertEqual(username,"admin")
        self.assertEqual(pwd,"hhhhhh")

if __name__ == '__main__':
    unittest.main()